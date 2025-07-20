import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def create_habit_tracker_excel(filename='habit_tracker.xlsx'):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = '주간 습관 성장 챌린지'

    # 상단 정보
    sheet['A1'] = '아이 이름:'
    sheet['B1'] = '___________________'
    sheet['D1'] = '주간 기간:'
    sheet['E1'] = '____년 __월 __일 ~ __일'
    sheet['G1'] = '이번 주 테마:'
    sheet['H1'] = '___________________'

    # 색상 코드 설명
    sheet['A3'] = '🎨 색상 코드'
    sheet['A4'] = '녹색 😊 = 목표 달성!'
    sheet['A5'] = '노랑 🤔 = 조금 아쉽지만 잘했어! (부분 달성)'
    sheet['A6'] = '빨강 😔 = 괜찮아, 내일 다시 해보자! (미달성)'

    # 헤더
    headers = ['시간대 (직접 목표를 적어보세요!)', '월요일', '화요일', '수요일', '목요일', '금요일', '토요일', '일요일', '주간 합계 (녹색 개수)']
    sheet.append(headers)

    # 헤더 스타일
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=8, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
        if col_idx > 1: # 요일과 주간 합계 열 너비 조정
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12
    sheet.column_dimensions['A'].width = 35 # 시간대 열 너비 조정

    # 예시 행 추가
    time_slots = [
        '아침 (예: 6-9시) 스스로 일어나기',
        '오전 (예: 9-12시) 집중해서 공부/놀이',
        '점심 (예: 12-1시) 편식 없이 골고루 먹기',
        '오후 (예: 1-5시) 스스로 계획한 일 하기',
        '저녁 (예: 6-9시) 정리 정돈 및 내일 준비',
        '(필요하면 자유롭게 추가하세요)'
    ]

    for i, slot in enumerate(time_slots):
        row_data = [slot] + ['[색칠]'] * 7 + ['/ 7']
        sheet.append(row_data)
        for col_idx in range(2, 9): # 색칠 칸에 테두리 추가
            cell = sheet.cell(row=9 + i, column=col_idx)
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # 주간 합계 셀에 테두리 추가
        sheet.cell(row=9 + i, column=9).border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
        sheet.cell(row=9 + i, column=9).alignment = Alignment(horizontal='center', vertical='center')

    # 돌아보기 섹션
    current_row = sheet.max_row + 2
    sheet.cell(row=current_row, column=1, value='📈 이번 주 돌아보기').font = Font(bold=True)
    current_row += 1
    sheet.cell(row=current_row, column=1, value='*   가장 초록색이 많았던 요일과 시간은 언제였나요?').font = Font(bold=True)
    sheet.cell(row=current_row, column=2, value='________________________________________________________')
    current_row += 1
    sheet.cell(row=current_row, column=1, value='*   어떤 습관이 가장 쉬웠고, 어떤 습관이 가장 어려웠나요?').font = Font(bold=True)
    sheet.cell(row=current_row, column=2, value='________________________________________________________')
    current_row += 1
    sheet.cell(row=current_row, column=1, value='*   다음 주에 더 잘하고 싶은 것은 무엇인가요?').font = Font(bold=True)
    sheet.cell(row=current_row, column=2, value='________________________________________________________')

    # 보상 아이디어 섹션
    current_row += 2
    sheet.cell(row=current_row, column=1, value='🏆 이번 주 보상 아이디어').font = Font(bold=True)
    current_row += 1
    sheet.cell(row=current_row, column=1, value='*   (아이와 함께 이번 주 목표 달성 시 받을 보상을 미리 정해보세요. 예: 영화 보기, 특별한 간식 먹기, 보드게임 시간 등)').font = Font(italic=True)
    sheet.cell(row=current_row, column=2, value='________________________________________________________')

    # 서명
    current_row += 2
    sheet.cell(row=current_row, column=1, value='서명: ___________________ (부모님)   ___________________ (나!)')

    workbook.save(filename)

if __name__ == '__main__':
    create_habit_tracker_excel()


